#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excelファイルの構造を分析するスクリプト
"""
import openpyxl
from pathlib import Path

def analyze_excel_file(file_path):
    """Excelファイルの構造を分析"""
    print(f"\n{'='*80}")
    print(f"ファイル: {file_path}")
    print(f"{'='*80}")

    wb = openpyxl.load_workbook(file_path, data_only=True)

    for sheet_name in wb.sheetnames:
        print(f"\n[シート名: {sheet_name}]")
        ws = wb[sheet_name]

        # ヘッダー行を表示（最初の2行）
        print("\nヘッダー行:")
        for row_idx in range(1, min(3, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(20, ws.max_column + 1)):  # 最初の20列まで
                cell_value = ws.cell(row_idx, col_idx).value
                row_data.append(f"[{col_idx}] {cell_value}")
            print(f"  行{row_idx}: {' | '.join(row_data)}")

        print(f"\n総行数: {ws.max_row}, 総列数: {ws.max_column}")

        # 所属情報関連のカラムを特定
        if ws.max_row > 0:
            header_row = []
            for col_idx in range(1, ws.max_column + 1):
                header_row.append(ws.cell(1, col_idx).value)

            # 所属情報関連のキーワードを検索
            keywords = ['section', 'group', 'project', 'tech', 'division', 'department', 'team',
                       '部', '課', '班', 'グループ', 'プロジェクト', '本部']

            print("\n所属情報関連カラム:")
            for idx, header in enumerate(header_row, 1):
                if header and any(keyword.lower() in str(header).lower() for keyword in keywords):
                    print(f"  列{idx}: {header}")

    wb.close()

if __name__ == "__main__":
    spreadsheet_dir = Path("/home/user/WorkEngagementSystem/SpreadSheet")

    # 主要なファイルを分析
    key_files = [
        "MemberSS.xlsx",
        "MasterSS.xlsx",
        "RatingSS.xlsx",
        "CommentSS.xlsx"
    ]

    for filename in key_files:
        file_path = spreadsheet_dir / filename
        if file_path.exists():
            try:
                analyze_excel_file(file_path)
            except Exception as e:
                print(f"\nエラー: {filename} の分析中にエラーが発生しました: {e}")
        else:
            print(f"\nファイルが見つかりません: {file_path}")
