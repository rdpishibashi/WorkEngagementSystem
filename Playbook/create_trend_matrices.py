#!/usr/bin/env python3
"""
trend_recent と trend_base の条件式マトリクス生成スクリプト
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# 定数値（we_analyzer.pyと同じ）
CHANGE_TAG_THRESHOLD = 6.0
TREND_RECENT_DELTA = 2.0
TREND_SLOPE = 0.5
TREND_SLOPE_STD_MIN = 0.2
TREND_SLOPE_STD = 0.45


def create_trend_recent_matrix():
    """trend_recent（短期トレンド）の条件マトリクスを作成"""

    data = [
        {
            'trend_recent': '連続上昇',
            'Priority': 1,
            '判定条件': 'E_delta_1 > TREND_RECENT_DELTA AND E_delta_1_prev > TREND_RECENT_DELTA',
            '条件式（数値）': 'E_delta_1 > 2.0 AND E_delta_1_prev > 2.0',
            '説明': '2期連続で先月比が+2.0超',
            'サンプル_E_delta_1': 3.5,
            'サンプル_E_delta_1_prev': 2.8,
        },
        {
            'trend_recent': '急上昇',
            'Priority': 2,
            '判定条件': 'E_delta_1 >= CHANGE_TAG_THRESHOLD',
            '条件式（数値）': 'E_delta_1 >= 6.0',
            '説明': '先月比が+6.0以上の大幅上昇',
            'サンプル_E_delta_1': 7.5,
            'サンプル_E_delta_1_prev': 1.2,
        },
        {
            'trend_recent': '上昇',
            'Priority': 3,
            '判定条件': 'TREND_RECENT_DELTA < E_delta_1 < CHANGE_TAG_THRESHOLD',
            '条件式（数値）': '2.0 < E_delta_1 < 6.0',
            '説明': '先月比が+2.0超~+6.0未満',
            'サンプル_E_delta_1': 3.5,
            'サンプル_E_delta_1_prev': 0.5,
        },
        {
            'trend_recent': '横ばい',
            'Priority': 4,
            '判定条件': '-TREND_RECENT_DELTA <= E_delta_1 <= TREND_RECENT_DELTA',
            '条件式（数値）': '-2.0 <= E_delta_1 <= 2.0',
            '説明': '先月比が-2.0~+2.0の範囲',
            'サンプル_E_delta_1': 0.5,
            'サンプル_E_delta_1_prev': -0.3,
        },
        {
            'trend_recent': '下降',
            'Priority': 5,
            '判定条件': '-CHANGE_TAG_THRESHOLD < E_delta_1 < -TREND_RECENT_DELTA',
            '条件式（数値）': '-6.0 < E_delta_1 < -2.0',
            '説明': '先月比が-6.0超~-2.0未満',
            'サンプル_E_delta_1': -3.5,
            'サンプル_E_delta_1_prev': -0.5,
        },
        {
            'trend_recent': '急落',
            'Priority': 6,
            '判定条件': 'E_delta_1 <= -CHANGE_TAG_THRESHOLD',
            '条件式（数値）': 'E_delta_1 <= -6.0',
            '説明': '先月比が-6.0以下の大幅低下',
            'サンプル_E_delta_1': -7.5,
            'サンプル_E_delta_1_prev': -1.2,
        },
        {
            'trend_recent': '連続下降',
            'Priority': 7,
            '判定条件': 'E_delta_1 < -TREND_RECENT_DELTA AND E_delta_1_prev < -TREND_RECENT_DELTA',
            '条件式（数値）': 'E_delta_1 < -2.0 AND E_delta_1_prev < -2.0',
            '説明': '2期連続で先月比が-2.0未満',
            'サンプル_E_delta_1': -3.5,
            'サンプル_E_delta_1_prev': -2.8,
        },
    ]

    return pd.DataFrame(data)


def create_trend_base_matrix():
    """trend_base（中期トレンド）の条件マトリクスを作成"""

    data = [
        {
            'trend_base': '未評価',
            'Priority': 1,
            '判定条件': '履歴不足（3ヶ月未満）',
            '条件式（数値）': 'データ件数 < 3',
            '説明': '中期トレンド計算に必要な履歴が不足',
            '使用指標': 'なし',
            'サンプル_E_slope_6': 'NaN',
            'サンプル_E_slope_6_std_12': 'NaN',
        },
        {
            'trend_base': '上昇中',
            'Priority': 2,
            '判定条件': '(E_slope_6 > 0 AND |E_slope_6| > TREND_SLOPE AND E_slope_6_std_12 > TREND_SLOPE_STD_MIN) OR (E_slope_6_std_12 > 0 AND E_slope_6_std_12 > TREND_SLOPE_STD)',
            '条件式（数値）': '(E_slope_6 > 0 AND |E_slope_6| > 0.5 AND E_slope_6_std_12 > 0.2) OR (E_slope_6_std_12 > 0 AND E_slope_6_std_12 > 0.45)',
            '説明': '6ヶ月傾き・標準化傾きが正で閾値超',
            '使用指標': 'E_slope_6, E_slope_6_std_12',
            'サンプル_E_slope_6': 0.65,
            'サンプル_E_slope_6_std_12': 0.55,
        },
        {
            'trend_base': '低下中',
            'Priority': 3,
            '判定条件': '(E_slope_6 < 0 AND |E_slope_6| > TREND_SLOPE AND E_slope_6_std_12 < -TREND_SLOPE_STD_MIN) OR (E_slope_6_std_12 < 0 AND E_slope_6_std_12 < -TREND_SLOPE_STD)',
            '条件式（数値）': '(E_slope_6 < 0 AND |E_slope_6| > 0.5 AND E_slope_6_std_12 < -0.2) OR (E_slope_6_std_12 < 0 AND E_slope_6_std_12 < -0.45)',
            '説明': '6ヶ月傾き・標準化傾きが負で閾値超',
            '使用指標': 'E_slope_6, E_slope_6_std_12',
            'サンプル_E_slope_6': -0.65,
            'サンプル_E_slope_6_std_12': -0.55,
        },
        {
            'trend_base': '安定',
            'Priority': 4,
            '判定条件': '上記以外（履歴あり、かつ上昇中・低下中の条件に該当しない）',
            '条件式（数値）': '|E_slope_6| <= 0.5 または |E_slope_6_std_12| <= 0.2/0.45',
            '説明': '傾きが小さく、明確なトレンドなし',
            '使用指標': 'E_slope_6, E_slope_6_std_12',
            'サンプル_E_slope_6': 0.15,
            'サンプル_E_slope_6_std_12': 0.12,
        },
    ]

    return pd.DataFrame(data)


def create_excel_with_matrices():
    """3つのマトリクスをExcelファイルに出力"""

    # trend_recentマトリクス
    trend_recent_df = create_trend_recent_matrix()

    # trend_baseマトリクス
    trend_base_df = create_trend_base_matrix()

    # trend_refined_condition.xlsxを読み込んで「横ばい」を「安定維持」に修正
    refined_df = pd.read_excel('trend_refined_condition.xlsx')
    # Priority 7の横ばいを安定維持に変更
    refined_df.loc[
        (refined_df['Priority'] == 7) &
        (refined_df['trend_refined'] == '横ばい'),
        'trend_refined'
    ] = '安定維持'

    # Excelファイルを作成
    wb = Workbook()

    # デフォルトシートを削除
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # 1. trend_recentマトリクス
    create_formatted_sheet(wb, 'trend_recent条件式', trend_recent_df,
                          "trend_recent（短期トレンド）の判定条件\n基準: E_delta_1（先月比）")

    # 2. trend_baseマトリクス
    create_formatted_sheet(wb, 'trend_base条件式', trend_base_df,
                          "trend_base（中期トレンド）の判定条件\n基準: E_slope_6（6ヶ月傾き）、E_slope_6_std_12（標準化傾き）")

    # 3. trend_refined条件（修正版）
    create_formatted_sheet(wb, 'trend_refined条件', refined_df,
                          "trend_refined（統合トレンド）の判定条件\nPriority順に評価")

    # 4. 定数一覧シート
    add_constants_sheet(wb)

    wb.save('trend_condition_matrices.xlsx')

    print("✓ トレンド条件マトリクスを作成しました: trend_condition_matrices.xlsx")
    print(f"  - trend_recent: {len(trend_recent_df)}種類")
    print(f"  - trend_base: {len(trend_base_df)}種類")
    print(f"  - trend_refined: {len(refined_df)}種類")


def create_formatted_sheet(wb, sheet_name, df, title):
    """書式設定付きでシートを作成"""

    ws = wb.create_sheet(sheet_name)

    # タイトル行
    ws.append([title])
    ws.merge_cells(f'A1:{chr(65+len(df.columns)-1)}1')
    title_cell = ws['A1']
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # 空行
    ws.append([])

    # データ
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 3):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # ヘッダー行
            if r_idx == 3:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

            # 枠線
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border

    # 列幅の自動調整
    for col_idx in range(1, len(df.columns) + 1):
        max_length = 0
        column_letter = chr(64 + col_idx)
        for row_idx in range(3, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width

    # ヘッダー行の高さ
    ws.row_dimensions[3].height = 30

    # 凍結ペイン
    ws.freeze_panes = 'A4'


def add_constants_sheet(wb):
    """定数一覧シートを追加"""

    ws = wb.create_sheet("定数一覧")

    constants_data = [
        ['定数名', '値', '説明'],
        ['', '', ''],
        ['CHANGE_TAG_THRESHOLD', 6.0, '急上昇・急落の閾値（E_delta_1の絶対値）'],
        ['TREND_RECENT_DELTA', 2.0, '上昇・下降の閾値（E_delta_1）'],
        ['', '', ''],
        ['TREND_SLOPE', 0.5, '上昇中・低下中の傾き閾値（E_slope_6の絶対値）'],
        ['TREND_SLOPE_STD_MIN', 0.2, '上昇中・低下中の標準化傾き最小閾値'],
        ['TREND_SLOPE_STD', 0.45, '上昇中・低下中の標準化傾き閾値'],
        ['', '', ''],
        ['BIG_CHANGE_PERSONAL_Z', 2.0, '個人内変化大の閾値（|E_delta_1|/E_std_12）'],
    ]

    for row in constants_data:
        ws.append(row)

    # 書式設定
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 50


if __name__ == "__main__":
    create_excel_with_matrices()
