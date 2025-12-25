#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MasterSS.xlsx と CommentSS.xlsx の組織階層を3階層から5階層に移行するスクリプト

変更内容:
旧6カラム構造:
  section | current_section | tech_group | current_tech_group | project_group | current_project_group

新10カラム構造:
  division | current_division | department | current_department | section | current_section | team | current_team | project | current_project

マッピング:
- division / current_division ← 空白
- department / current_department ← 旧 section / current_section
- section / current_section ← 旧 tech_group / current_tech_group
- team / current_team ← 空白
- project / current_project ← 旧 project_group / current_project_group
"""
import openpyxl
from pathlib import Path
import shutil
from datetime import datetime

def backup_file(file_path):
    """ファイルのバックアップを作成"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = file_path.parent / f"{file_path.stem}_backup_{timestamp}{file_path.suffix}"
    shutil.copy2(file_path, backup_path)
    print(f"バックアップ作成: {backup_path}")
    return backup_path

def migrate_sheet_6col_to_10col(ws, sheet_name, start_col):
    """
    6カラム構造を10カラム構造に変更

    Args:
        ws: ワークシート
        sheet_name: シート名（ログ用）
        start_col: 所属情報の開始列（section の列番号）
    """
    print(f"\n  [{sheet_name}] 移行開始...")
    print(f"    開始列: {start_col}")
    print(f"    移行前の列数: {ws.max_column}")
    print(f"    移行前の行数: {ws.max_row}")

    # 列の位置
    OLD_SECTION_COL = start_col
    OLD_CURRENT_SECTION_COL = start_col + 1
    OLD_TECH_GROUP_COL = start_col + 2
    OLD_CURRENT_TECH_GROUP_COL = start_col + 3
    OLD_PROJECT_GROUP_COL = start_col + 4
    OLD_CURRENT_PROJECT_GROUP_COL = start_col + 5

    NEW_DIVISION_COL = start_col
    NEW_CURRENT_DIVISION_COL = start_col + 1
    NEW_DEPARTMENT_COL = start_col + 2
    NEW_CURRENT_DEPARTMENT_COL = start_col + 3
    NEW_SECTION_COL = start_col + 4
    NEW_CURRENT_SECTION_COL = start_col + 5
    NEW_TEAM_COL = start_col + 6
    NEW_CURRENT_TEAM_COL = start_col + 7
    NEW_PROJECT_COL = start_col + 8
    NEW_CURRENT_PROJECT_COL = start_col + 9

    # 6カラム後の列を4列右にシフト
    shift_start_col = OLD_CURRENT_PROJECT_GROUP_COL + 1
    max_col = ws.max_column

    if max_col >= shift_start_col:
        print(f"    列シフト処理開始 (列{shift_start_col}以降を4列右へ)...")
        for row_idx in range(1, ws.max_row + 1):
            # 後ろから前に向かってコピー
            for col_idx in range(max_col, shift_start_col - 1, -1):
                old_cell = ws.cell(row_idx, col_idx)
                new_cell = ws.cell(row_idx, col_idx + 4)

                new_cell.value = old_cell.value

                if old_cell.has_style:
                    new_cell.font = old_cell.font.copy()
                    new_cell.border = old_cell.border.copy()
                    new_cell.fill = old_cell.fill.copy()
                    new_cell.number_format = old_cell.number_format
                    new_cell.protection = old_cell.protection.copy()
                    new_cell.alignment = old_cell.alignment.copy()
        print(f"    列シフト完了")

    # ヘッダー行を更新
    print(f"    ヘッダー行を更新...")
    ws.cell(1, NEW_DIVISION_COL).value = "division"
    ws.cell(1, NEW_CURRENT_DIVISION_COL).value = "current_division"
    ws.cell(1, NEW_DEPARTMENT_COL).value = "department"
    ws.cell(1, NEW_CURRENT_DEPARTMENT_COL).value = "current_department"
    ws.cell(1, NEW_SECTION_COL).value = "section"
    ws.cell(1, NEW_CURRENT_SECTION_COL).value = "current_section"
    ws.cell(1, NEW_TEAM_COL).value = "team"
    ws.cell(1, NEW_CURRENT_TEAM_COL).value = "current_team"
    ws.cell(1, NEW_PROJECT_COL).value = "project"
    ws.cell(1, NEW_CURRENT_PROJECT_COL).value = "current_project"

    # データ行を移行
    print(f"    データ移行開始...")
    for row_idx in range(2, ws.max_row + 1):
        # 旧データを取得
        old_section = ws.cell(row_idx, OLD_SECTION_COL).value
        old_current_section = ws.cell(row_idx, OLD_CURRENT_SECTION_COL).value
        old_tech_group = ws.cell(row_idx, OLD_TECH_GROUP_COL).value
        old_current_tech_group = ws.cell(row_idx, OLD_CURRENT_TECH_GROUP_COL).value
        old_project_group = ws.cell(row_idx, OLD_PROJECT_GROUP_COL).value
        old_current_project_group = ws.cell(row_idx, OLD_CURRENT_PROJECT_GROUP_COL).value

        # 新しい列に設定
        ws.cell(row_idx, NEW_DIVISION_COL).value = ""
        ws.cell(row_idx, NEW_CURRENT_DIVISION_COL).value = ""
        ws.cell(row_idx, NEW_DEPARTMENT_COL).value = old_section
        ws.cell(row_idx, NEW_CURRENT_DEPARTMENT_COL).value = old_current_section
        ws.cell(row_idx, NEW_SECTION_COL).value = old_tech_group
        ws.cell(row_idx, NEW_CURRENT_SECTION_COL).value = old_current_tech_group
        ws.cell(row_idx, NEW_TEAM_COL).value = ""
        ws.cell(row_idx, NEW_CURRENT_TEAM_COL).value = ""
        ws.cell(row_idx, NEW_PROJECT_COL).value = old_project_group
        ws.cell(row_idx, NEW_CURRENT_PROJECT_COL).value = old_current_project_group

    # 列幅を調整
    for col in [NEW_DIVISION_COL, NEW_CURRENT_DIVISION_COL, NEW_TEAM_COL, NEW_CURRENT_TEAM_COL]:
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    for col in [NEW_DEPARTMENT_COL, NEW_CURRENT_DEPARTMENT_COL, NEW_SECTION_COL, NEW_CURRENT_SECTION_COL, NEW_PROJECT_COL, NEW_CURRENT_PROJECT_COL]:
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    print(f"    データ移行完了")
    print(f"    移行後の列数: {ws.max_column}")

def migrate_sheet_3col_to_5col(ws, sheet_name, start_col):
    """
    3カラム構造を5カラム構造に変更（rating2, rating3シート用）

    Args:
        ws: ワークシート
        sheet_name: シート名（ログ用）
        start_col: 所属情報の開始列（section の列番号）
    """
    print(f"\n  [{sheet_name}] 移行開始...")
    print(f"    開始列: {start_col}")
    print(f"    移行前の列数: {ws.max_column}")
    print(f"    移行前の行数: {ws.max_row}")

    # 列の位置
    OLD_SECTION_COL = start_col
    OLD_GROUP_COL = start_col + 1
    OLD_PROJECT_COL = start_col + 2

    NEW_DIVISION_COL = start_col
    NEW_DEPARTMENT_COL = start_col + 1
    NEW_SECTION_COL = start_col + 2
    NEW_TEAM_COL = start_col + 3
    NEW_PROJECT_COL = start_col + 4

    # 3カラム後の列を2列右にシフト
    shift_start_col = OLD_PROJECT_COL + 1
    max_col = ws.max_column

    if max_col >= shift_start_col:
        print(f"    列シフト処理開始 (列{shift_start_col}以降を2列右へ)...")
        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(max_col, shift_start_col - 1, -1):
                old_cell = ws.cell(row_idx, col_idx)
                new_cell = ws.cell(row_idx, col_idx + 2)

                new_cell.value = old_cell.value

                if old_cell.has_style:
                    new_cell.font = old_cell.font.copy()
                    new_cell.border = old_cell.border.copy()
                    new_cell.fill = old_cell.fill.copy()
                    new_cell.number_format = old_cell.number_format
                    new_cell.protection = old_cell.protection.copy()
                    new_cell.alignment = old_cell.alignment.copy()
        print(f"    列シフト完了")

    # ヘッダー行を更新
    print(f"    ヘッダー行を更新...")
    ws.cell(1, NEW_DIVISION_COL).value = "division"
    ws.cell(1, NEW_DEPARTMENT_COL).value = "department"
    ws.cell(1, NEW_SECTION_COL).value = "section"
    ws.cell(1, NEW_TEAM_COL).value = "team"
    ws.cell(1, NEW_PROJECT_COL).value = "project"

    # データ行を移行
    print(f"    データ移行開始...")
    for row_idx in range(2, ws.max_row + 1):
        # 旧データを取得
        old_section = ws.cell(row_idx, OLD_SECTION_COL).value
        old_group = ws.cell(row_idx, OLD_GROUP_COL).value
        old_project = ws.cell(row_idx, OLD_PROJECT_COL).value

        # 新しい列に設定
        ws.cell(row_idx, NEW_DIVISION_COL).value = ""
        ws.cell(row_idx, NEW_DEPARTMENT_COL).value = old_section
        ws.cell(row_idx, NEW_SECTION_COL).value = old_group
        ws.cell(row_idx, NEW_TEAM_COL).value = ""
        ws.cell(row_idx, NEW_PROJECT_COL).value = old_project

    # 列幅を調整
    ws.column_dimensions[openpyxl.utils.get_column_letter(NEW_DIVISION_COL)].width = 15
    ws.column_dimensions[openpyxl.utils.get_column_letter(NEW_DEPARTMENT_COL)].width = 20
    ws.column_dimensions[openpyxl.utils.get_column_letter(NEW_SECTION_COL)].width = 20
    ws.column_dimensions[openpyxl.utils.get_column_letter(NEW_TEAM_COL)].width = 15
    ws.column_dimensions[openpyxl.utils.get_column_letter(NEW_PROJECT_COL)].width = 20

    print(f"    データ移行完了")
    print(f"    移行後の列数: {ws.max_column}")

def migrate_master_ss(file_path):
    """MasterSS.xlsxを移行"""
    print(f"\n{'='*80}")
    print(f"MasterSS.xlsx の移行を開始します")
    print(f"{'='*80}")

    backup_file(file_path)

    wb = openpyxl.load_workbook(file_path)

    # rating シート: 6カラム構造、開始列7
    if 'rating' in wb.sheetnames:
        migrate_sheet_6col_to_10col(wb['rating'], 'rating', 7)

    # rating2 シート: 3カラム構造、開始列6
    if 'rating2' in wb.sheetnames:
        migrate_sheet_3col_to_5col(wb['rating2'], 'rating2', 6)

    # rating3 シート: 3カラム構造、開始列6
    if 'rating3' in wb.sheetnames:
        migrate_sheet_3col_to_5col(wb['rating3'], 'rating3', 6)

    # evaluation シート: 6カラム構造、開始列7
    if 'evaluation' in wb.sheetnames:
        migrate_sheet_6col_to_10col(wb['evaluation'], 'evaluation', 7)

    # comment シート: 6カラム構造、開始列7
    if 'comment' in wb.sheetnames:
        migrate_sheet_6col_to_10col(wb['comment'], 'comment', 7)

    wb.save(file_path)
    wb.close()

    print(f"\n✓ MasterSS.xlsx の移行が完了しました")
    print(f"{'='*80}")

def migrate_comment_ss(file_path):
    """CommentSS.xlsxを移行"""
    print(f"\n{'='*80}")
    print(f"CommentSS.xlsx の移行を開始します")
    print(f"{'='*80}")

    backup_file(file_path)

    wb = openpyxl.load_workbook(file_path)

    # comments シート: 6カラム構造、開始列7
    if 'comments' in wb.sheetnames:
        migrate_sheet_6col_to_10col(wb['comments'], 'comments', 7)

    wb.save(file_path)
    wb.close()

    print(f"\n✓ CommentSS.xlsx の移行が完了しました")
    print(f"{'='*80}")

if __name__ == "__main__":
    # MasterSS.xlsx を移行
    master_path = Path("/home/user/WorkEngagementSystem/SpreadSheet/MasterSS.xlsx")
    if master_path.exists():
        migrate_master_ss(master_path)
    else:
        print(f"エラー: ファイルが見つかりません: {master_path}")

    # CommentSS.xlsx を移行
    comment_path = Path("/home/user/WorkEngagementSystem/SpreadSheet/CommentSS.xlsx")
    if comment_path.exists():
        migrate_comment_ss(comment_path)
    else:
        print(f"エラー: ファイルが見つかりません: {comment_path}")

    print("\n" + "="*80)
    print("全てのExcelファイルの移行が完了しました")
    print("="*80)
