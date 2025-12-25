#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MemberSS.xlsx の組織階層を3階層から5階層に移行するスクリプト

変更内容:
- 旧: section (列6) | tech_group (列7) | project_group (列8)
- 新: division (列6) | department (列7) | section (列8) | team (列9) | project (列10)

マッピング:
- division ← 空白（後で手作業で編集）
- department ← 旧 section
- section ← 旧 tech_group
- team ← 空白（後で手作業で編集）
- project ← 旧 project_group
"""
import openpyxl
from pathlib import Path
import shutil
from datetime import datetime

def backup_file(file_path):
    """ファイルのバックアップを作成"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = file_path.parent / f"{file_path.stem}_backup_{timestamp}{file_path.suffix}"
    shutil.copy2(file_path, backup_path)
    print(f"バックアップ作成: {backup_path}")
    return backup_path

def migrate_member_ss(file_path):
    """MemberSS.xlsxを移行"""
    print(f"\n{'='*80}")
    print(f"MemberSS.xlsx の移行を開始します")
    print(f"{'='*80}")

    # バックアップ作成
    backup_file(file_path)

    # ファイルを開く
    wb = openpyxl.load_workbook(file_path)
    ws = wb['members']

    print(f"\n移行前の列数: {ws.max_column}")
    print(f"移行前の行数: {ws.max_row}")

    # 列の位置
    OLD_SECTION_COL = 6
    OLD_TECH_GROUP_COL = 7
    OLD_PROJECT_GROUP_COL = 8

    NEW_DIVISION_COL = 6
    NEW_DEPARTMENT_COL = 7
    NEW_SECTION_COL = 8
    NEW_TEAM_COL = 9
    NEW_PROJECT_COL = 10

    # 列9以降を2列右にシフト（後ろから処理）
    # 最大列数を確認
    max_col = ws.max_column

    print(f"\n列シフト処理開始...")
    # 各行について処理
    for row_idx in range(1, ws.max_row + 1):
        # 後ろから前に向かってコピー（上書きを防ぐため）
        for col_idx in range(max_col, OLD_PROJECT_GROUP_COL, -1):
            old_cell = ws.cell(row_idx, col_idx)
            new_cell = ws.cell(row_idx, col_idx + 2)

            # 値をコピー
            new_cell.value = old_cell.value

            # スタイルをコピー
            if old_cell.has_style:
                new_cell.font = old_cell.font.copy()
                new_cell.border = old_cell.border.copy()
                new_cell.fill = old_cell.fill.copy()
                new_cell.number_format = old_cell.number_format
                new_cell.protection = old_cell.protection.copy()
                new_cell.alignment = old_cell.alignment.copy()

    print(f"列シフト完了")

    # ヘッダー行を更新（行1）
    print(f"\nヘッダー行を更新...")
    ws.cell(1, NEW_DIVISION_COL).value = "division"
    ws.cell(1, NEW_DEPARTMENT_COL).value = "department"
    ws.cell(1, NEW_SECTION_COL).value = "section"
    ws.cell(1, NEW_TEAM_COL).value = "team"
    ws.cell(1, NEW_PROJECT_COL).value = "project"

    # データ行を移行（行2以降）
    print(f"データ移行開始...")
    for row_idx in range(2, ws.max_row + 1):
        # 旧データを取得
        old_section = ws.cell(row_idx, OLD_SECTION_COL).value
        old_tech_group = ws.cell(row_idx, OLD_TECH_GROUP_COL).value
        old_project_group = ws.cell(row_idx, OLD_PROJECT_GROUP_COL).value

        # 新しい列に設定
        ws.cell(row_idx, NEW_DIVISION_COL).value = ""  # 空白
        ws.cell(row_idx, NEW_DEPARTMENT_COL).value = old_section
        ws.cell(row_idx, NEW_SECTION_COL).value = old_tech_group
        ws.cell(row_idx, NEW_TEAM_COL).value = ""  # 空白
        ws.cell(row_idx, NEW_PROJECT_COL).value = old_project_group

    print(f"データ移行完了")

    # 列幅を調整
    ws.column_dimensions[openpyxl.utils.get_column_letter(NEW_DIVISION_COL)].width = 15
    ws.column_dimensions[openpyxl.utils.get_column_letter(NEW_DEPARTMENT_COL)].width = 20
    ws.column_dimensions[openpyxl.utils.get_column_letter(NEW_SECTION_COL)].width = 20
    ws.column_dimensions[openpyxl.utils.get_column_letter(NEW_TEAM_COL)].width = 15
    ws.column_dimensions[openpyxl.utils.get_column_letter(NEW_PROJECT_COL)].width = 20

    print(f"\n移行後の列数: {ws.max_column}")

    # 保存
    wb.save(file_path)
    wb.close()

    print(f"\n✓ MemberSS.xlsx の移行が完了しました")
    print(f"{'='*80}")

if __name__ == "__main__":
    file_path = Path("/home/user/WorkEngagementSystem/SpreadSheet/MemberSS.xlsx")

    if not file_path.exists():
        print(f"エラー: ファイルが見つかりません: {file_path}")
        exit(1)

    migrate_member_ss(file_path)

    print("\n移行後の確認:")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb['members']

    # ヘッダー行を表示
    print("\nヘッダー行 (列1-15):")
    for col_idx in range(1, 16):
        print(f"  列{col_idx}: {ws.cell(1, col_idx).value}")

    # サンプルデータを表示（行2）
    if ws.max_row >= 2:
        print("\nサンプルデータ (行2, 列1-15):")
        for col_idx in range(1, 16):
            print(f"  列{col_idx}: {ws.cell(2, col_idx).value}")

    wb.close()
